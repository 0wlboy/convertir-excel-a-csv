import csv 
from openpyxl import load_workbook 

wb = load_workbook(filename='archivo.xlsx')
hoja = wb.active 

csv_datos = []
for valor in hoja.iter_rows(values_only = True):
    csv_datos.append(list(valor))

with open('archivo.csv', 'w') as csv_obj:
    escritor = csv.writer(csv_obj,delimiter='.')
    for line in csv_datos:
        escritor.writerow(line)